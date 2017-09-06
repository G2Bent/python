from Heygears.Data import handing_data
from selenium import webdriver
from openpyxl.styles import Font
from openpyxl.styles.colors import RED,BLUE
import openpyxl,time
class hg_excel(handing_data):
    #初始化url
    def __init__(self,url,file_url):
        self.test_url = url
        self.file_url = file_url

    def Verify_Excel(self,filename):
        #读取表格的sheet列表集合
        workbook = openpyxl.load_workbook(self.file_url)
        sheets = workbook.get_sheet_names()

        #Operation操作方法
        Operation_list = ['click','double_click','move','drag_drop',
                          'send','send_code','clear','Back_Space',
                          'iframe','get_attribute','assert_text','assert_url',
                          'get_current_url','upfile','cookie','forward','back',
                          'assert_attribute']

        #不需要填写element的Operation操作
        Operation_element_is_empty = [
            'cookie','get_current_url','assert_url','forward','back'
        ]

        #method的方法
        element_method = [
            'xpath','css','text','partial_text',
            'tag_name','name','class_name','id'
        ]

        True_count = 0      # 用于记录True的数量
        result  = list()    # 返回的测试结果数列

        # 对每个表格进行验证
        for sheet_name in sheets:
            sheet = workbook.get_sheet_by_name(sheet_name)
            #行和列的总数
            rows = len(list(sheet.rows))
            columns = len(list(sheet.columns))

            file_content = list()
            for x in range(2,rows+1):
                row_list = list()
                for y in range(1,columns+1):
                    row_list.append(sheet.cell(row=x,column=y).value)
                file_content.append(row_list)

            #将内容转化为字典
            content_dict = handing_data.file_content_to_dict(self,file_content)
            #进行校验
            for (key,dicts) in content_dict.items():
                True_count ,state= 0, 'False' # 用于记录True的数量，初始化每一行校验检验的状态
                #判断操作是否存在，在操作存在的情况下，判断页面定位是否合法
                if dicts['Operation'] != None:
                    if dicts['Operation'] not in Operation_list:
                        print ('文件--<'+filename+'>--中的表格--<'+sheet_name+'>--第'+str(key+1)+'行，不存在'+dicts['Operation']+'这个操作！')
                    else:
                        if dicts['Operation'] not in Operation_element_is_empty:
                            if dicts['element'] == None or dicts['element'] == '':
                                print ('文件--<'+filename+'>--中的表格--<'+sheet_name+'>--第'+str(key+1)+'行，页面定位不能为空！')
                            else:
                                element_str = dicts['element'].split(',')
                                if len(element_str) is not 2:
                                    print ('文件--<'+filename+'>--中的表格--<'+sheet_name+'>--第'+str(key+1)+'行，页面定位-请按照（method,path）的方法填写')
                                else:
                                    method,xpath = element_str[0],element_str[1]
                                    if method not in element_method:
                                        print('文件--<' + filename + '>--中的表格--<'+sheet_name+'>--第' + str(key+1) + '行，页面定位-（method,path）的method不存在')
                                    elif xpath == None or xpath == '':
                                        print ('文件--<' + filename + '>--中的表格--<'+sheet_name+'>--第' + str(key+1) + '行，页面定位-（method,path）的xpath不能为空')
                                    else:
                                        True_count += 1
                        else:
                            if dicts['element'] != None:
                                print ('文件--<'+filename+'>--中的表格--<'+sheet_name+'>--第'+str(key+1)+'行，操作'+dicts['Operation']+'页面定位必须为空！')
                            else:
                                True_count += 1
                else:
                    if dicts['element'] == None:
                        True_count += 1
                    else:
                        print ('文件--<'+filename+'>--中的表格--<'+sheet_name+'>--第'+str(key+1)+'行，操作没有值的时候，页面定位不能有值!')


                #判断截图是否规范
                if dicts['picture'] != None:
                    a_list = dicts['picture'].split(',')
                    if len(a_list) != 2:
                        print ('文件--<' + filename + '>--中的表格--<'+sheet_name+'>--第' + str(key+1) + '行，截图的正确填写格式为：(Whether_picture,scrollTop)')
                    else:
                        Whether_picture, scrollTop = a_list[0], a_list[1]
                        if int(scrollTop)<0:
                            print ('文件--<' + filename + '>--中的表格--<'+sheet_name+'>--第' + str(key+1) + '行，截图逗号右边填写的必须是正整数！')
                        else:
                            if Whether_picture == 'Yes' or Whether_picture == 'yes' or Whether_picture == '是':
                                True_count += 1
                            else:
                                print ('文件--<' + filename + '>--中的表格--<'+sheet_name+'>--第' + str(key+1) + '行，截图逗号左边填写的必须是 Yes 或者 yes 或者 是！')
                else:
                    True_count += 1

                #判断输出定位是否符合规范
                if dicts['Output_element'] !=None:
                    element_str = dicts['Output_element'].split(',')
                    if len(element_str) is not 2:
                        print('文件--<' + filename + '>--中的表格--<'+sheet_name+'>--第' + str(key+1) + '行，输出定位-请按照（method,path）的方法填写')
                    else:
                        method, xpath = element_str[0], element_str[1]
                        if method not in element_method:
                            print('文件--<' + filename + '>--中的表格--<'+sheet_name+'>--第' + str(key+1) + '行，输出定位-（method,path）的method不存在')
                        elif xpath == None or xpath == '':
                            print('文件--<' + filename + '>--中的表格--<'+sheet_name+'>--第' + str(key+1) + '行，输出定位-（method,path）的xpath不能为空')
                        else:
                            True_count += 1
                else:
                    True_count += 1


                #判断时间
                if dicts['wait_time'] != None:
                    if int(dicts['wait_time']) < 0:
                        print ('文件--<' + filename + '>--中的表格--<'+sheet_name+'>--第' + str(key+1) + '行，等待时间-请输入正整数！')
                    else:
                        True_count += 1
                else:
                    True_count += 1

                #如果四个校验都通过，就添加一个true，（一个true代表的是一个通过）
                if True_count == 4 :
                    state = 'True'
                result.append(state)
        return result




    #将excel转化为selenium进行测试
    def Open_Excel(self,file_name,Report_picture_url,report_sheet):
        #读取表格的sheet列表集合
        workbook = openpyxl.load_workbook(self.file_url)
        sheets = workbook.get_sheet_names()

        #单个文件通过个数
        pass_count, fail_count, test_result= 0, 0, list()

        #对每个表格进行操作
        index = 2
        picture_index = 0
        for sheet_name in sheets:
            driver = webdriver.Chrome()
            driver.get(self.test_url)
            driver.maximize_window()
            sheet = workbook.get_sheet_by_name(sheet_name)
            report_sheet['A%d' % (index) ] = sheet_name

            row, column, file_content = len(list(sheet.rows)), len(list(sheet.columns)),[]
            #读取每一页的内容，生成一个总列表
            for x in range(2,row+1):
                row_list = []
                for y in range(1,column+1):
                    row_list.append(sheet.cell(row=x,column=y).value)
                file_content.append(row_list)
            #把内容转化为字典，以便判断
            hd = handing_data(driver=driver)
            content_dict = hd.file_content_to_dict(file_content)
            try:
                a = hd.test_hg_selenium(content_dict,Report_picture_url,report_sheet,index)
                output,picture_index = a[0],a[1]
                pass_count += 1
                #写入设置返回的值
                report_sheet['B%d' % (index)] = str(output)
                #写入是否通过，通过为蓝色的字体，不通过为红色的字体
                report_sheet['C%d' % (index)] = 'Pass'
                report_sheet['C%d' % (index)].font = Font(color=BLUE)
                print(sheet_name,index)
            except Exception as e:
                fail_count +=1
                #写入不通过的异常
                report_sheet['B%d' % (index)] = str(e)
                print (sheet_name,index,str(e))
                # 写入是否通过，通过为蓝色的字体，不通过为红色的字体
                report_sheet['C%d' % (index)] = 'Fail'
                report_sheet['C%d' % (index)].font = Font(color=RED)
                picture_index += 1
            finally:
                driver.close()
            # print ('picture_index:'+str(picture_index))
            index = picture_index
            index +=1
            # print ('index:'+str(index))
            # print ('------')

        test_result.append(pass_count)
        test_result.append(fail_count)
        return test_result